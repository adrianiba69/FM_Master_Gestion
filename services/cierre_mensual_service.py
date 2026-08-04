import shutil
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from database import conectar
from pdf.resumen_pdf import ResumenPDF
from runtime_paths import CIERRES_DIR
from services.backup_service import BackupService
from services.resumen_service import ResumenService
from services.servicio_service import ServicioService


class CierreMensualService:
    """Orquesta el cierre mensual sin repetir períodos ya resumidos."""

    def __init__(self, fecha=None):
        self.fecha = self._fecha(fecha or date.today())
        self.periodo = self.fecha.strftime("%Y-%m")
        self.carpeta = CIERRES_DIR / self.periodo
        self.carpeta_resumenes = self.carpeta / "resumenes"
        self.errores = []
        self.resumenes_generados = []
        self.renovaciones = []
        self.backup_path = None
        self.carpeta_resumenes.mkdir(parents=True, exist_ok=True)

    def analizar_servicios(self):
        ServicioService.actualizar_estados_periodo(self.fecha)
        conn = conectar()
        cur = conn.cursor()
        cur.execute("""
            SELECT s.id,
                   COALESCE(NULLIF(c.razon_social, ''), c.nombre),
                   s.concepto, s.fecha_inicio, s.fecha_fin,
                   COALESCE(s.renovable, 1),
                   (COALESCE(s.cantidad, 0) * COALESCE(s.importe, 0))
                       - COALESCE(s.descuento, 0)
            FROM servicios s
            JOIN clientes c ON c.id=s.cliente_id
            WHERE s.activo=1 AND s.fecha_fin<=?
            ORDER BY s.fecha_fin, 2, s.concepto
        """, ((self.fecha + timedelta(days=7)).isoformat(),))
        filas = cur.fetchall()
        conn.close()

        resultado = {"vencidos": [], "hoy": [], "proximos": []}
        for fila in filas:
            fin = self._fecha(fila[4])
            item = {
                "id": fila[0], "cliente": fila[1], "concepto": fila[2],
                "fecha_inicio": fila[3], "fecha_fin": fila[4],
                "renovable": bool(fila[5]), "importe": float(fila[6] or 0),
            }
            if fin < self.fecha:
                resultado["vencidos"].append(item)
            elif fin == self.fecha:
                resultado["hoy"].append(item)
            else:
                resultado["proximos"].append(item)
        return resultado

    def clientes_pendientes(self):
        return ResumenService.listar_pendientes(self.fecha)

    def generar_resumenes(self, cliente_ids):
        resultado = {"generados": 0, "omitidos": 0, "importe_total": 0.0, "errores": []}
        for cliente_id in dict.fromkeys(int(valor) for valor in cliente_ids):
            resumen = None
            try:
                resumen = ResumenService.generar_pendiente_cliente(
                    cliente_id, fecha=self.fecha
                )
                if resumen is None:
                    resultado["omitidos"] += 1
                    continue
                ruta_generada = ResumenPDF.generar(resumen.id)
                registro = {
                    "id": resumen.id,
                    "numero": resumen.numero,
                    "cliente_id": resumen.cliente_id,
                    "total": float(resumen.total or 0),
                    "ruta": str(Path(ruta_generada).resolve()),
                }
                self.resumenes_generados.append(registro)
                resultado["generados"] += 1
                resultado["importe_total"] += registro["total"]
            except Exception as error:
                if resumen is not None:
                    try:
                        ResumenService.eliminar_generacion(resumen.id)
                    except Exception as error_limpieza:
                        self.errores.append(f"Limpieza de resumen {resumen.id}: {error_limpieza}")
                mensaje = f"Cliente {cliente_id}: {error}"
                resultado["errores"].append(mensaje)
                self.errores.append(mensaje)
        return resultado

    def renovar_servicios(self, renovar):
        if not renovar:
            return {"renovados": [], "omitidos": 0, "errores": []}
        analisis = self.analizar_servicios()
        ids = [
            item["id"]
            for grupo in (analisis["vencidos"], analisis["hoy"])
            for item in grupo
            if item["renovable"]
        ]
        resultado = ServicioService.renovar_periodos(ids)
        self.renovaciones.extend(resultado["renovados"])
        self.errores.extend(resultado["errores"])
        return resultado

    def crear_backup(self):
        ruta_original = Path(BackupService.crear_backup())
        destino = self._ruta_disponible(self.carpeta / ruta_original.name)
        shutil.copy2(ruta_original, destino)
        self.backup_path = str(destino.resolve())
        return self.backup_path

    def exportar_informes(self):
        datos = self._datos_informe()
        ruta_pdf = self.carpeta / f"informe_cierre_{self.periodo}.pdf"
        ruta_excel = self.carpeta / f"informe_cierre_{self.periodo}.xlsx"
        self._exportar_pdf(datos, ruta_pdf)
        self._exportar_excel(datos, ruta_excel)
        return {"pdf": str(ruta_pdf.resolve()), "excel": str(ruta_excel.resolve())}

    def _datos_informe(self):
        inicio = self.fecha.replace(day=1).isoformat()
        if self.fecha.month == 12:
            siguiente = self.fecha.replace(year=self.fecha.year + 1, month=1, day=1)
        else:
            siguiente = self.fecha.replace(month=self.fecha.month + 1, day=1)
        conn = conectar()
        cur = conn.cursor()
        cur.execute("""
            SELECT r.numero, r.fecha,
                   COALESCE(NULLIF(c.razon_social, ''), c.nombre),
                   r.total, r.saldo, r.estado
            FROM resumenes r JOIN clientes c ON c.id=r.cliente_id
            WHERE r.fecha>=? AND r.fecha<?
            ORDER BY r.numero
        """, (inicio, siguiente.isoformat()))
        filas = cur.fetchall()
        conn.close()
        return {
            "titulo": f"Cierre mensual {self.periodo}",
            "columnas": ("Resumen", "Fecha", "Cliente", "Total", "Saldo", "Estado"),
            "filas": filas,
            "total": sum(float(fila[3] or 0) for fila in filas),
            "generados": len(self.resumenes_generados),
            "renovados": len(self.renovaciones),
            "errores": list(self.errores),
            "backup": self.backup_path or "No generado",
        }

    @classmethod
    def _exportar_pdf(cls, datos, ruta):
        documento = SimpleDocTemplate(
            str(ruta), pagesize=landscape(A4), leftMargin=12 * mm,
            rightMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm,
            title=datos["titulo"],
        )
        estilos = getSampleStyleSheet()
        elementos = [Paragraph(datos["titulo"].upper(), estilos["Title"]), Spacer(1, 4 * mm)]
        resumen = (
            f"Resúmenes generados en el asistente: {datos['generados']} &nbsp;&nbsp; "
            f"Importe mensual: {cls.moneda(datos['total'])} &nbsp;&nbsp; "
            f"Renovaciones: {datos['renovados']} &nbsp;&nbsp; Errores: {len(datos['errores'])}"
        )
        elementos.extend([Paragraph(resumen, estilos["BodyText"]), Spacer(1, 4 * mm)])
        tabla_datos = [list(datos["columnas"])]
        tabla_datos.extend([
            [fila[0], fila[1], fila[2], cls.moneda(fila[3]), cls.moneda(fila[4]), fila[5]]
            for fila in datos["filas"]
        ])
        if len(tabla_datos) == 1:
            tabla_datos.append(["Sin resúmenes", "", "", "", "", ""])
        tabla = Table(tabla_datos, repeatRows=1)
        tabla.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#C00000")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#BBBBBB")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F3F3")]),
            ("ALIGN", (3, 1), (4, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        elementos.append(tabla)
        if datos["errores"]:
            elementos.extend([Spacer(1, 4 * mm), Paragraph("Errores", estilos["Heading3"])])
            elementos.extend(Paragraph(str(error), estilos["BodyText"]) for error in datos["errores"])
        documento.build(elementos)

    @classmethod
    def _exportar_excel(cls, datos, ruta):
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Cierre mensual"
        hoja.merge_cells("A1:F1")
        hoja["A1"] = datos["titulo"].upper()
        hoja["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        hoja["A1"].fill = PatternFill("solid", fgColor="C00000")
        hoja["A1"].alignment = Alignment(horizontal="center")
        hoja.append([])
        hoja.append(["Generados", datos["generados"], "Importe mensual", datos["total"], "Renovaciones", datos["renovados"]])
        hoja.append([])
        hoja.append(list(datos["columnas"]))
        for celda in hoja[5]:
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = PatternFill("solid", fgColor="1B1B1B")
        for fila in datos["filas"]:
            hoja.append(list(fila))
        for fila in range(6, hoja.max_row + 1):
            hoja.cell(fila, 4).number_format = '$ #,##0.00'
            hoja.cell(fila, 5).number_format = '$ #,##0.00'
        hoja.column_dimensions["C"].width = 38
        for indice in (1, 2, 4, 5, 6):
            hoja.column_dimensions[get_column_letter(indice)].width = 17
        hoja.freeze_panes = "A6"
        hoja_errores = libro.create_sheet("Errores")
        hoja_errores.append(["Detalle"])
        for error in datos["errores"]:
            hoja_errores.append([str(error)])
        hoja_errores.column_dimensions["A"].width = 100
        libro.save(ruta)

    @staticmethod
    def moneda(valor):
        numero = f"{float(valor or 0):,.2f}"
        return "$ " + numero.replace(",", "X").replace(".", ",").replace("X", ".")

    @staticmethod
    def _fecha(valor):
        if isinstance(valor, datetime):
            return valor.date()
        if isinstance(valor, date):
            return valor
        return datetime.strptime(str(valor), "%Y-%m-%d").date()

    @staticmethod
    def _ruta_disponible(ruta):
        if not ruta.exists():
            return ruta
        marca = datetime.now().strftime("%H%M%S")
        return ruta.with_name(f"{ruta.stem}_{marca}{ruta.suffix}")
