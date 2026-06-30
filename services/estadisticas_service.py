from collections import defaultdict
from datetime import date, datetime
from io import BytesIO

from matplotlib.backends.backend_agg import FigureCanvasAgg
from matplotlib.figure import Figure
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from database import conectar
from runtime_paths import EXPORTS_DIR


class EstadisticasService:
    EXPORTS_DIR = EXPORTS_DIR / "estadisticas"

    @classmethod
    def obtener(cls, desde=None, hasta=None, cliente_id=None, servicio="Todos", estado="Todos"):
        desde = cls._fecha(desde)
        hasta = cls._fecha(hasta)
        if desde and hasta and desde > hasta:
            raise ValueError("La fecha desde no puede ser posterior a la fecha hasta.")
        cliente_id = int(cliente_id) if cliente_id else None
        servicio = None if servicio in (None, "", "Todos") else servicio
        estado = None if estado in (None, "", "Todos") else estado
        conn = conectar()
        try:
            facturacion = cls._facturacion_mensual(conn, desde, hasta, cliente_id, servicio)
            cobros = cls._cobros_mensuales(conn, desde, hasta, cliente_id)
            saldo = cls._saldo_pendiente(conn, desde, hasta, cliente_id, servicio)
            clientes_activos = cls._clientes_activos(conn, cliente_id)
            servicios_activos = cls._servicios_activos(conn, cliente_id, servicio)
            oportunidades = cls._oportunidades(conn, desde, hasta, cliente_id, estado)
            ranking_clientes = cls._ranking_clientes(conn, desde, hasta, cliente_id, servicio)
            ranking_servicios = cls._ranking_servicios(conn, desde, hasta, cliente_id, servicio)
            deuda = cls._evolucion_deuda(conn, desde, hasta, cliente_id, servicio)
        finally:
            conn.close()
        return {
            "indicadores": {
                "facturacion_mensual": sum(valor for _, valor in facturacion),
                "cobros_mensuales": sum(valor for _, valor in cobros),
                "saldo_pendiente": saldo,
                "clientes_activos": clientes_activos,
                "servicios_activos": servicios_activos,
                "oportunidades_abiertas": oportunidades["abiertas"],
                "oportunidades_ganadas": oportunidades["ganadas"],
            },
            "facturacion_mensual": facturacion,
            "cobros_mensuales": cobros,
            "oportunidades_estado": oportunidades["estados"],
            "ranking_clientes": ranking_clientes,
            "ranking_servicios": ranking_servicios,
            "deuda_mensual": deuda,
            "filtros": {
                "desde": desde or "", "hasta": hasta or "",
                "cliente_id": cliente_id, "servicio": servicio or "Todos",
                "estado": estado or "Todos",
            },
        }

    @staticmethod
    def listar_clientes():
        conn = conectar()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, COALESCE(NULLIF(razon_social, ''), nombre)
            FROM clientes ORDER BY 2
        """)
        filas = cur.fetchall()
        conn.close()
        return filas

    @staticmethod
    def listar_servicios():
        conn = conectar()
        cur = conn.cursor()
        cur.execute("""
            SELECT DISTINCT concepto FROM servicios
            WHERE TRIM(COALESCE(concepto, ''))!='' ORDER BY concepto
        """)
        filas = [fila[0] for fila in cur.fetchall()]
        conn.close()
        return filas

    @classmethod
    def exportar_excel(cls, datos):
        ruta = cls._ruta("xlsx")
        libro = Workbook()
        resumen = libro.active
        resumen.title = "Resumen"
        resumen.append(["ESTADÍSTICAS COMERCIALES"])
        resumen.merge_cells("A1:B1")
        cls._titulo_excel(resumen["A1"])
        nombres = {
            "facturacion_mensual": "Facturación",
            "cobros_mensuales": "Cobros",
            "saldo_pendiente": "Saldo pendiente",
            "clientes_activos": "Clientes activos",
            "servicios_activos": "Servicios activos",
            "oportunidades_abiertas": "Oportunidades abiertas",
            "oportunidades_ganadas": "Oportunidades ganadas",
        }
        resumen.append([])
        resumen.append(["Indicador", "Valor"])
        for clave, valor in datos["indicadores"].items():
            resumen.append([nombres[clave], valor])
        for fila in range(4, 11):
            if fila <= 6:
                resumen.cell(fila, 2).number_format = '$ #,##0.00'
        cls._hoja_series(libro, "Facturación mensual", ("Mes", "Facturación"), datos["facturacion_mensual"], moneda_columna=2)
        cls._hoja_series(libro, "Cobros mensuales", ("Mes", "Cobros"), datos["cobros_mensuales"], moneda_columna=2)
        cls._hoja_series(libro, "Oportunidades", ("Estado", "Cantidad"), datos["oportunidades_estado"])
        cls._hoja_series(libro, "Ranking clientes", ("Cliente", "Facturación"), datos["ranking_clientes"], moneda_columna=2)
        cls._hoja_series(libro, "Ranking servicios", ("Servicio", "Cantidad"), datos["ranking_servicios"])
        cls._hoja_series(libro, "Evolución deuda", ("Mes", "Saldo acumulado"), datos["deuda_mensual"], moneda_columna=2)
        for hoja in libro.worksheets:
            for columna in range(1, hoja.max_column + 1):
                ancho = max(len(str(hoja.cell(fila, columna).value or "")) for fila in range(1, hoja.max_row + 1))
                hoja.column_dimensions[get_column_letter(columna)].width = min(max(12, ancho + 3), 42)
        libro.save(ruta)
        return str(ruta)

    @classmethod
    def exportar_pdf(cls, datos):
        ruta = cls._ruta("pdf")
        documento = SimpleDocTemplate(
            str(ruta), pagesize=landscape(A4), leftMargin=10 * mm,
            rightMargin=10 * mm, topMargin=10 * mm, bottomMargin=10 * mm,
            title="Estadísticas comerciales",
        )
        estilos = getSampleStyleSheet()
        elementos = [Paragraph("ESTADÍSTICAS COMERCIALES", estilos["Title"]), Spacer(1, 4 * mm)]
        indicadores = datos["indicadores"]
        tabla_resumen = [
            ["Facturación", "Cobros", "Saldo", "Clientes activos", "Servicios activos", "Op. abiertas", "Op. ganadas"],
            [
                cls.formatear_moneda(indicadores["facturacion_mensual"]),
                cls.formatear_moneda(indicadores["cobros_mensuales"]),
                cls.formatear_moneda(indicadores["saldo_pendiente"]),
                indicadores["clientes_activos"], indicadores["servicios_activos"],
                indicadores["oportunidades_abiertas"], indicadores["oportunidades_ganadas"],
            ],
        ]
        tabla = Table(tabla_resumen, repeatRows=1)
        tabla.setStyle(cls._estilo_tabla())
        elementos.extend([tabla, Spacer(1, 5 * mm)])
        imagen = cls._imagen_graficos(datos)
        elementos.append(Image(imagen, width=260 * mm, height=118 * mm))
        elementos.extend([Spacer(1, 4 * mm), Paragraph("RANKINGS", estilos["Heading2"])])
        rankings = [["Cliente", "Facturación", "Servicio", "Cantidad"]]
        maximo = max(len(datos["ranking_clientes"]), len(datos["ranking_servicios"]), 1)
        for indice in range(maximo):
            cliente = datos["ranking_clientes"][indice] if indice < len(datos["ranking_clientes"]) else ("", "")
            servicio = datos["ranking_servicios"][indice] if indice < len(datos["ranking_servicios"]) else ("", "")
            rankings.append([cliente[0], cls.formatear_moneda(cliente[1]) if cliente[0] else "", servicio[0], servicio[1]])
        tabla_ranking = Table(rankings, colWidths=[75 * mm, 35 * mm, 75 * mm, 30 * mm], repeatRows=1)
        tabla_ranking.setStyle(cls._estilo_tabla())
        elementos.append(tabla_ranking)
        documento.build(elementos)
        return str(ruta)

    @staticmethod
    def _facturacion_mensual(conn, desde, hasta, cliente_id, servicio):
        condiciones, parametros = [], []
        if desde:
            condiciones.append("r.fecha>=?"); parametros.append(desde)
        if hasta:
            condiciones.append("r.fecha<=?"); parametros.append(hasta)
        if cliente_id:
            condiciones.append("r.cliente_id=?"); parametros.append(cliente_id)
        if servicio:
            consulta = """
                SELECT SUBSTR(r.fecha,1,7), COALESCE(SUM(rc.total),0)
                FROM resumen_conceptos rc JOIN resumenes r ON r.id=rc.resumen_id
                WHERE rc.concepto=?
            """
            parametros.insert(0, servicio)
        else:
            consulta = "SELECT SUBSTR(r.fecha,1,7), COALESCE(SUM(r.total),0) FROM resumenes r"
            if condiciones:
                consulta += " WHERE "
        if servicio:
            if condiciones:
                consulta += " AND " + " AND ".join(condiciones)
        elif condiciones:
            consulta += " AND ".join(condiciones)
        consulta += " GROUP BY 1 ORDER BY 1"
        cur = conn.cursor(); cur.execute(consulta, parametros)
        return [(fila[0], float(fila[1] or 0)) for fila in cur.fetchall() if fila[0]]

    @staticmethod
    def _cobros_mensuales(conn, desde, hasta, cliente_id):
        condiciones, parametros = [], []
        if desde: condiciones.append("fecha>=?"); parametros.append(desde)
        if hasta: condiciones.append("fecha<=?"); parametros.append(hasta)
        if cliente_id: condiciones.append("cliente_id=?"); parametros.append(cliente_id)
        consulta = "SELECT SUBSTR(fecha,1,7), COALESCE(SUM(importe),0) FROM cobros"
        if condiciones: consulta += " WHERE " + " AND ".join(condiciones)
        consulta += " GROUP BY 1 ORDER BY 1"
        cur = conn.cursor(); cur.execute(consulta, parametros)
        return [(fila[0], float(fila[1] or 0)) for fila in cur.fetchall() if fila[0]]

    @staticmethod
    def _saldo_pendiente(conn, desde, hasta, cliente_id, servicio):
        condiciones, parametros = ["r.saldo>0"], []
        if desde: condiciones.append("r.fecha>=?"); parametros.append(desde)
        if hasta: condiciones.append("r.fecha<=?"); parametros.append(hasta)
        if cliente_id: condiciones.append("r.cliente_id=?"); parametros.append(cliente_id)
        if servicio:
            condiciones.append("EXISTS(SELECT 1 FROM resumen_conceptos rc WHERE rc.resumen_id=r.id AND rc.concepto=?)")
            parametros.append(servicio)
        cur = conn.cursor(); cur.execute("SELECT COALESCE(SUM(r.saldo),0) FROM resumenes r WHERE " + " AND ".join(condiciones), parametros)
        return float(cur.fetchone()[0] or 0)

    @staticmethod
    def _clientes_activos(conn, cliente_id):
        consulta = "SELECT COUNT(*) FROM clientes WHERE LOWER(TRIM(COALESCE(estado,'')))='activo'"
        parametros = []
        if cliente_id: consulta += " AND id=?"; parametros.append(cliente_id)
        cur = conn.cursor(); cur.execute(consulta, parametros)
        return cur.fetchone()[0]

    @staticmethod
    def _servicios_activos(conn, cliente_id, servicio):
        condiciones, parametros = ["activo=1"] , []
        if cliente_id: condiciones.append("cliente_id=?"); parametros.append(cliente_id)
        if servicio: condiciones.append("concepto=?"); parametros.append(servicio)
        cur = conn.cursor(); cur.execute("SELECT COUNT(*) FROM servicios WHERE " + " AND ".join(condiciones), parametros)
        return cur.fetchone()[0]

    @staticmethod
    def _oportunidades(conn, desde, hasta, cliente_id, estado):
        condiciones, parametros = [], []
        if desde: condiciones.append("fecha>=?"); parametros.append(desde)
        if hasta: condiciones.append("fecha<=?"); parametros.append(hasta)
        if cliente_id: condiciones.append("cliente_id=?"); parametros.append(cliente_id)
        if estado: condiciones.append("estado=?"); parametros.append(estado)
        donde = " WHERE " + " AND ".join(condiciones) if condiciones else ""
        cur = conn.cursor(); cur.execute("SELECT estado, COUNT(*) FROM oportunidades" + donde + " GROUP BY estado ORDER BY estado", parametros)
        estados = [(fila[0] or "Sin estado", fila[1]) for fila in cur.fetchall()]
        mapa = dict(estados)
        return {
            "estados": estados,
            "abiertas": sum(valor for clave, valor in estados if clave not in ("Ganada", "Perdida")),
            "ganadas": mapa.get("Ganada", 0),
        }

    @staticmethod
    def _ranking_clientes(conn, desde, hasta, cliente_id, servicio):
        condiciones, parametros = [], []
        if desde: condiciones.append("r.fecha>=?"); parametros.append(desde)
        if hasta: condiciones.append("r.fecha<=?"); parametros.append(hasta)
        if cliente_id: condiciones.append("r.cliente_id=?"); parametros.append(cliente_id)
        if servicio:
            condiciones.append("EXISTS(SELECT 1 FROM resumen_conceptos rc WHERE rc.resumen_id=r.id AND rc.concepto=?)")
            parametros.append(servicio)
        consulta = """
            SELECT COALESCE(NULLIF(c.razon_social,''), c.nombre), COALESCE(SUM(r.total),0)
            FROM resumenes r JOIN clientes c ON c.id=r.cliente_id
        """
        if condiciones: consulta += " WHERE " + " AND ".join(condiciones)
        consulta += " GROUP BY r.cliente_id ORDER BY 2 DESC LIMIT 10"
        cur = conn.cursor(); cur.execute(consulta, parametros)
        return [(fila[0], float(fila[1] or 0)) for fila in cur.fetchall()]

    @staticmethod
    def _ranking_servicios(conn, desde, hasta, cliente_id, servicio):
        condiciones, parametros = [], []
        if desde: condiciones.append("r.fecha>=?"); parametros.append(desde)
        if hasta: condiciones.append("r.fecha<=?"); parametros.append(hasta)
        if cliente_id: condiciones.append("r.cliente_id=?"); parametros.append(cliente_id)
        if servicio: condiciones.append("rc.concepto=?"); parametros.append(servicio)
        consulta = """
            SELECT rc.concepto, COALESCE(SUM(rc.cantidad),0)
            FROM resumen_conceptos rc JOIN resumenes r ON r.id=rc.resumen_id
        """
        if condiciones: consulta += " WHERE " + " AND ".join(condiciones)
        consulta += " GROUP BY rc.concepto ORDER BY 2 DESC LIMIT 10"
        cur = conn.cursor(); cur.execute(consulta, parametros)
        return [(fila[0] or "Sin concepto", float(fila[1] or 0)) for fila in cur.fetchall()]

    @classmethod
    def _evolucion_deuda(cls, conn, desde, hasta, cliente_id, servicio):
        facturacion = dict(cls._facturacion_mensual(conn, desde, hasta, cliente_id, servicio))
        cobros = dict(cls._cobros_mensuales(conn, desde, hasta, cliente_id))
        meses = sorted(set(facturacion) | set(cobros))
        acumulado, resultado = 0, []
        for mes in meses:
            acumulado += facturacion.get(mes, 0) - cobros.get(mes, 0)
            resultado.append((mes, max(0, acumulado)))
        return resultado

    @classmethod
    def _imagen_graficos(cls, datos):
        figura = Figure(figsize=(13, 5.9), dpi=120, tight_layout=True)
        ejes = [figura.add_subplot(2, 3, indice) for indice in range(1, 6)]
        cls._barras(ejes[0], datos["facturacion_mensual"], "Facturación por mes", "#C00000")
        cls._barras(ejes[1], datos["cobros_mensuales"], "Cobros por mes", "#1B1B1B")
        estados = datos["oportunidades_estado"]
        if estados:
            ejes[2].pie([x[1] for x in estados], labels=[x[0] for x in estados], autopct="%1.0f%%", textprops={"fontsize": 7})
        else: ejes[2].text(.5, .5, "Sin datos", ha="center", va="center")
        ejes[2].set_title("Estado de oportunidades")
        cls._barras(ejes[3], datos["ranking_clientes"], "Ranking de clientes", "#C00000", horizontal=True)
        serie = datos["deuda_mensual"]
        if serie:
            ejes[4].plot([x[0] for x in serie], [x[1] for x in serie], marker="o", color="#C00000")
            ejes[4].tick_params(axis="x", rotation=45, labelsize=7)
        else: ejes[4].text(.5, .5, "Sin datos", ha="center", va="center")
        ejes[4].set_title("Evolución de saldo pendiente")
        salida = BytesIO(); FigureCanvasAgg(figura).print_png(salida); salida.seek(0)
        return salida

    @staticmethod
    def _barras(eje, serie, titulo, color, horizontal=False):
        if serie:
            etiquetas, valores = zip(*serie)
            if horizontal:
                eje.barh(etiquetas, valores, color=color)
            else:
                eje.bar(etiquetas, valores, color=color)
                eje.tick_params(axis="x", rotation=45, labelsize=7)
        else:
            eje.text(.5, .5, "Sin datos", ha="center", va="center")
        eje.set_title(titulo)

    @classmethod
    def _ruta(cls, extension):
        cls.EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
        marca = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        return cls.EXPORTS_DIR / f"estadisticas_{marca}.{extension}"

    @staticmethod
    def _titulo_excel(celda):
        celda.font = Font(size=16, bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="C00000")
        celda.alignment = Alignment(horizontal="center")

    @classmethod
    def _hoja_series(cls, libro, titulo, columnas, filas, moneda_columna=None):
        hoja = libro.create_sheet(titulo[:31])
        hoja.append(list(columnas))
        for celda in hoja[1]:
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = PatternFill("solid", fgColor="1B1B1B")
        for fila in filas: hoja.append(list(fila))
        if moneda_columna:
            for indice in range(2, hoja.max_row + 1):
                hoja.cell(indice, moneda_columna).number_format = '$ #,##0.00'

    @staticmethod
    def _estilo_tabla():
        return TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#C00000")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), .4, colors.HexColor("#BDBDBD")),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ])

    @staticmethod
    def _fecha(valor):
        if not valor: return None
        if isinstance(valor, date): return valor.isoformat()
        return datetime.strptime(str(valor), "%Y-%m-%d").date().isoformat()

    @staticmethod
    def formatear_moneda(valor):
        numero = f"{float(valor or 0):,.2f}"
        return "$ " + numero.replace(",", "X").replace(".", ",").replace("X", ".")
