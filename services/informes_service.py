import re
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from database import conectar
from runtime_paths import EXPORTS_DIR


class InformesService:
    INFORMES = (
        "Clientes activos",
        "Clientes con saldo pendiente",
        "Servicios activos",
        "Servicios próximos a vencer",
        "Resúmenes emitidos por período",
        "Cobros por período",
        "Facturación mensual",
        "Deuda total por cliente",
    )
    ESTADOS = (
        "Todos",
        "Activo",
        "Inactivo",
        "Pendiente",
        "Pagado",
        "Vencido",
        "Finalizado",
    )
    EXPORTS_DIR = EXPORTS_DIR
    COLUMNAS_MONEDA = {"Importe", "Total", "Facturado", "Cobrado", "Saldo", "Deuda"}

    @classmethod
    def generar(cls, informe, fecha_desde=None, fecha_hasta=None, cliente_id=None, estado=None):
        if informe not in cls.INFORMES:
            raise ValueError("Seleccione un informe válido.")
        desde = cls._fecha(fecha_desde)
        hasta = cls._fecha(fecha_hasta)
        if desde and hasta and desde > hasta:
            raise ValueError("La fecha desde no puede ser posterior a la fecha hasta.")
        estado = None if estado in (None, "", "Todos") else estado

        metodos = {
            "Clientes activos": cls._clientes_activos,
            "Clientes con saldo pendiente": cls._clientes_saldo,
            "Servicios activos": cls._servicios_activos,
            "Servicios próximos a vencer": cls._servicios_vencer,
            "Resúmenes emitidos por período": cls._resumenes_periodo,
            "Cobros por período": cls._cobros_periodo,
            "Facturación mensual": cls._facturacion_mensual,
            "Deuda total por cliente": cls._deuda_clientes,
        }
        columnas, filas = metodos[informe](desde, hasta, cliente_id, estado)
        return {
            "titulo": informe,
            "columnas": columnas,
            "filas": filas,
            "filtros": {
                "desde": desde or "",
                "hasta": hasta or "",
                "cliente_id": cliente_id,
                "estado": estado or "Todos",
            },
        }

    @staticmethod
    def _consultar(consulta, parametros=()):
        conn = conectar()
        cur = conn.cursor()
        cur.execute(consulta, parametros)
        filas = cur.fetchall()
        conn.close()
        return filas

    @classmethod
    def _clientes_activos(cls, _desde, _hasta, cliente_id, estado):
        condiciones = ["LOWER(TRIM(COALESCE(c.estado, '')))='activo'"]
        parametros = []
        if cliente_id:
            condiciones.append("c.id=?")
            parametros.append(cliente_id)
        if estado:
            condiciones.append("LOWER(c.estado)=LOWER(?)")
            parametros.append(estado)
        filas = cls._consultar(f"""
            SELECT c.id, c.codigo,
                   COALESCE(NULLIF(c.razon_social, ''), c.nombre),
                   c.telefono, c.whatsapp, c.localidad, c.estado
            FROM clientes c
            WHERE {' AND '.join(condiciones)}
            ORDER BY 3
        """, parametros)
        return ("ID", "Código", "Cliente", "Teléfono", "WhatsApp", "Localidad", "Estado"), filas

    @classmethod
    def _clientes_saldo(cls, _desde, _hasta, cliente_id, estado):
        condiciones = ["(COALESCE(r.facturado, 0)-COALESCE(co.cobrado, 0))>0"]
        parametros = []
        if cliente_id:
            condiciones.append("c.id=?")
            parametros.append(cliente_id)
        if estado in ("Activo", "Inactivo"):
            condiciones.append("LOWER(c.estado)=LOWER(?)")
            parametros.append(estado)
        filas = cls._consultar(f"""
            SELECT c.id,
                   COALESCE(NULLIF(c.razon_social, ''), c.nombre),
                   COALESCE(r.facturado, 0), COALESCE(co.cobrado, 0),
                   COALESCE(r.facturado, 0)-COALESCE(co.cobrado, 0), c.estado
            FROM clientes c
            LEFT JOIN (
                SELECT cliente_id, SUM(total) facturado FROM resumenes GROUP BY cliente_id
            ) r ON r.cliente_id=c.id
            LEFT JOIN (
                SELECT cliente_id, SUM(importe) cobrado FROM cobros GROUP BY cliente_id
            ) co ON co.cliente_id=c.id
            WHERE {' AND '.join(condiciones)}
            ORDER BY 5 DESC, 2
        """, parametros)
        return ("ID", "Cliente", "Facturado", "Cobrado", "Saldo", "Estado"), filas

    @classmethod
    def _servicios_activos(cls, desde, hasta, cliente_id, estado):
        hoy = date.today().isoformat()
        condiciones = ["s.activo=1", "s.fecha_inicio<=?", "s.fecha_fin>=?"]
        parametros = [hasta or hoy, desde or hoy]
        if cliente_id:
            condiciones.append("s.cliente_id=?")
            parametros.append(cliente_id)
        if estado in ("Activo", "Vencido", "Finalizado"):
            condiciones.append("LOWER(s.estado_periodo)=LOWER(?)")
            parametros.append(estado)
        filas = cls._consultar(f"""
            SELECT COALESCE(NULLIF(c.razon_social, ''), c.nombre),
                   s.concepto, s.fecha_inicio, s.fecha_fin, s.cantidad,
                   s.importe, s.descuento,
                   (s.cantidad*s.importe)-s.descuento, s.estado_periodo
            FROM servicios s JOIN clientes c ON c.id=s.cliente_id
            WHERE {' AND '.join(condiciones)}
            ORDER BY 1, s.concepto
        """, parametros)
        return (
            "Cliente", "Concepto", "Fecha inicio", "Fecha fin", "Cantidad",
            "Importe", "Descuento", "Total", "Estado",
        ), filas

    @classmethod
    def _servicios_vencer(cls, desde, hasta, cliente_id, estado):
        inicio = desde or date.today().isoformat()
        fin = hasta or (date.today() + timedelta(days=7)).isoformat()
        condiciones = ["s.activo=1", "s.fecha_fin BETWEEN ? AND ?"]
        parametros = [inicio, fin]
        if cliente_id:
            condiciones.append("s.cliente_id=?")
            parametros.append(cliente_id)
        if estado in ("Activo", "Vencido", "Finalizado"):
            condiciones.append("LOWER(s.estado_periodo)=LOWER(?)")
            parametros.append(estado)
        filas = cls._consultar(f"""
            SELECT COALESCE(NULLIF(c.razon_social, ''), c.nombre),
                   s.concepto, s.fecha_inicio, s.fecha_fin, s.importe,
                   s.renovable, s.estado_periodo
            FROM servicios s JOIN clientes c ON c.id=s.cliente_id
            WHERE {' AND '.join(condiciones)}
            ORDER BY s.fecha_fin, 1
        """, parametros)
        return ("Cliente", "Concepto", "Fecha inicio", "Fecha fin", "Importe", "Renovable", "Estado"), filas

    @classmethod
    def _resumenes_periodo(cls, desde, hasta, cliente_id, estado):
        condiciones, parametros = [], []
        cls._agregar_periodo(condiciones, parametros, "r.fecha", desde, hasta)
        if cliente_id:
            condiciones.append("r.cliente_id=?")
            parametros.append(cliente_id)
        if estado:
            condiciones.append("LOWER(r.estado)=LOWER(?)")
            parametros.append(estado)
        where = " WHERE " + " AND ".join(condiciones) if condiciones else ""
        filas = cls._consultar(f"""
            SELECT r.numero, r.fecha, r.fecha_vencimiento,
                   COALESCE(NULLIF(c.razon_social, ''), c.nombre),
                   r.total, r.saldo, r.estado
            FROM resumenes r JOIN clientes c ON c.id=r.cliente_id
            {where} ORDER BY r.fecha DESC, r.numero DESC
        """, parametros)
        return ("Resumen", "Fecha", "Vencimiento", "Cliente", "Total", "Saldo", "Estado"), filas

    @classmethod
    def _cobros_periodo(cls, desde, hasta, cliente_id, _estado):
        condiciones, parametros = [], []
        cls._agregar_periodo(condiciones, parametros, "co.fecha", desde, hasta)
        if cliente_id:
            condiciones.append("co.cliente_id=?")
            parametros.append(cliente_id)
        where = " WHERE " + " AND ".join(condiciones) if condiciones else ""
        filas = cls._consultar(f"""
            SELECT co.fecha, COALESCE(NULLIF(c.razon_social, ''), c.nombre),
                   co.importe, co.forma_pago, co.comprobante, co.observaciones
            FROM cobros co JOIN clientes c ON c.id=co.cliente_id
            {where} ORDER BY co.fecha DESC, co.id DESC
        """, parametros)
        return ("Fecha", "Cliente", "Importe", "Forma de pago", "Comprobante", "Observaciones"), filas

    @classmethod
    def _facturacion_mensual(cls, desde, hasta, cliente_id, estado):
        condiciones, parametros = [], []
        cls._agregar_periodo(condiciones, parametros, "r.fecha", desde, hasta)
        if cliente_id:
            condiciones.append("r.cliente_id=?")
            parametros.append(cliente_id)
        if estado:
            condiciones.append("LOWER(r.estado)=LOWER(?)")
            parametros.append(estado)
        where = " WHERE " + " AND ".join(condiciones) if condiciones else ""
        filas = cls._consultar(f"""
            SELECT SUBSTR(r.fecha, 1, 7), COUNT(*), SUM(r.total), SUM(r.saldo)
            FROM resumenes r
            {where}
            GROUP BY SUBSTR(r.fecha, 1, 7)
            ORDER BY 1 DESC
        """, parametros)
        return ("Mes", "Resúmenes", "Facturado", "Saldo"), filas

    @classmethod
    def _deuda_clientes(cls, _desde, _hasta, cliente_id, estado):
        condiciones, parametros = [], []
        if cliente_id:
            condiciones.append("c.id=?")
            parametros.append(cliente_id)
        if estado in ("Activo", "Inactivo"):
            condiciones.append("LOWER(c.estado)=LOWER(?)")
            parametros.append(estado)
        where = " WHERE " + " AND ".join(condiciones) if condiciones else ""
        filas = cls._consultar(f"""
            SELECT c.id, COALESCE(NULLIF(c.razon_social, ''), c.nombre),
                   COALESCE(r.facturado, 0), COALESCE(co.cobrado, 0),
                   COALESCE(r.facturado, 0)-COALESCE(co.cobrado, 0), c.estado
            FROM clientes c
            LEFT JOIN (
                SELECT cliente_id, SUM(total) facturado FROM resumenes GROUP BY cliente_id
            ) r ON r.cliente_id=c.id
            LEFT JOIN (
                SELECT cliente_id, SUM(importe) cobrado FROM cobros GROUP BY cliente_id
            ) co ON co.cliente_id=c.id
            {where} ORDER BY 5 DESC, 2
        """, parametros)
        return ("ID", "Cliente", "Facturado", "Cobrado", "Deuda", "Estado"), filas

    @staticmethod
    def _agregar_periodo(condiciones, parametros, campo, desde, hasta):
        if desde:
            condiciones.append(f"{campo}>=?")
            parametros.append(desde)
        if hasta:
            condiciones.append(f"{campo}<=?")
            parametros.append(hasta)

    @classmethod
    def exportar_excel(cls, datos):
        ruta = cls._ruta_exportacion(datos["titulo"], ".xlsx")
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Informe"
        columnas = datos["columnas"]
        ultima_columna = get_column_letter(max(1, len(columnas)))
        hoja.merge_cells(f"A1:{ultima_columna}1")
        titulo = hoja["A1"]
        titulo.value = datos["titulo"].upper()
        titulo.font = Font(size=16, bold=True, color="FFFFFF")
        titulo.fill = PatternFill("solid", fgColor="C00000")
        titulo.alignment = Alignment(horizontal="center")
        hoja.row_dimensions[1].height = 28

        for indice, columna in enumerate(columnas, start=1):
            celda = hoja.cell(row=3, column=indice, value=columna)
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = PatternFill("solid", fgColor="1B1B1B")
            celda.alignment = Alignment(horizontal="center")
        for fila_indice, fila in enumerate(datos["filas"], start=4):
            for columna_indice, valor in enumerate(fila, start=1):
                celda = hoja.cell(row=fila_indice, column=columna_indice, value=valor)
                if columnas[columna_indice - 1] in cls.COLUMNAS_MONEDA:
                    celda.number_format = '$ #,##0.00'
        for indice, columna in enumerate(columnas, start=1):
            maximo = max(
                [len(str(columna)), *[
                    len(str(fila[indice - 1] or "")) for fila in datos["filas"]
                ]]
            )
            hoja.column_dimensions[get_column_letter(indice)].width = min(maximo + 3, 45)
        hoja.freeze_panes = "A4"
        hoja.auto_filter.ref = f"A3:{ultima_columna}{max(3, len(datos['filas']) + 3)}"
        libro.save(ruta)
        return str(ruta)

    @classmethod
    def exportar_pdf(cls, datos):
        ruta = cls._ruta_exportacion(datos["titulo"], ".pdf")
        documento = SimpleDocTemplate(
            str(ruta),
            pagesize=landscape(A4),
            leftMargin=12 * mm,
            rightMargin=12 * mm,
            topMargin=12 * mm,
            bottomMargin=12 * mm,
            title=datos["titulo"],
        )
        estilos = getSampleStyleSheet()
        elementos = [Paragraph(datos["titulo"].upper(), estilos["Title"]), Spacer(1, 6 * mm)]
        estilo_celda = estilos["BodyText"]
        estilo_celda.fontSize = 7
        estilo_celda.leading = 9
        tabla_datos = [[Paragraph(str(columna), estilo_celda) for columna in datos["columnas"]]]
        for fila in datos["filas"]:
            tabla_datos.append([
                Paragraph(cls._texto_pdf(valor, datos["columnas"][indice]), estilo_celda)
                for indice, valor in enumerate(fila)
            ])
        if len(tabla_datos) == 1:
            tabla_datos.append([Paragraph("Sin resultados", estilo_celda)] + [""] * (len(datos["columnas"]) - 1))
        ancho_disponible = landscape(A4)[0] - 24 * mm
        tabla = Table(
            tabla_datos,
            colWidths=[ancho_disponible / len(datos["columnas"])] * len(datos["columnas"]),
            repeatRows=1,
        )
        tabla.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#C00000")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#BDBDBD")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F4F4")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]))
        elementos.append(tabla)
        documento.build(elementos)
        return str(ruta)

    @classmethod
    def _ruta_exportacion(cls, titulo, extension):
        cls.EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
        normalizado = unicodedata.normalize("NFKD", titulo).encode("ascii", "ignore").decode()
        slug = re.sub(r"[^a-z0-9]+", "_", normalizado.lower()).strip("_")
        marca = datetime.now().strftime("%Y%m%d_%H%M%S")
        ruta = cls.EXPORTS_DIR / f"{slug}_{marca}{extension}"
        contador = 1
        while ruta.exists():
            ruta = cls.EXPORTS_DIR / f"{slug}_{marca}_{contador:02d}{extension}"
            contador += 1
        return ruta

    @staticmethod
    def _fecha(valor):
        if not valor:
            return None
        if isinstance(valor, date):
            return valor.isoformat()
        return datetime.strptime(str(valor), "%Y-%m-%d").date().isoformat()

    @classmethod
    def _texto_pdf(cls, valor, columna):
        if valor is None:
            return ""
        if columna in cls.COLUMNAS_MONEDA:
            return cls.formatear_moneda(valor)
        if columna == "Renovable":
            return "Sí" if valor else "No"
        return str(valor)

    @staticmethod
    def formatear_moneda(valor):
        numero = f"{float(valor or 0):,.2f}"
        return "$ " + numero.replace(",", "X").replace(".", ",").replace("X", ".")
